#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Validador de Expediente de Reclutamiento — Fitness Para Todos
================================================================

Lee un conjunto de PDFs (uno por documento), identifica de qué documento se
trata, extrae texto (con OCR automático si el PDF es una imagen escaneada),
corre las reglas de validación que pidió el equipo de reclutamiento, y
genera un Excel (.xlsx) con el resultado para anexar al expediente.

REGLAS IMPLEMENTADAS
---------------------
1. Legibilidad: se extrae texto nativo del PDF; si una página no trae texto
   (documento escaneado) se corre OCR (Tesseract, español). Si ni el texto
   nativo ni el OCR logran leer nada útil, se marca "ILEGIBLE".
2. Pertenencia: se compara el nombre encontrado en cada documento contra el
   nombre registrado del candidato (comparación por tokens, tolerante a
   orden Nombre/Apellido y a acentos). El Comprobante de domicilio se
   excluye de esta regla (puede estar a nombre de otra persona).
3. Completitud: se coteja contra la lista de documentos obligatorios y se
   reporta qué falta. Cada documento debe llegar en un solo PDF (esto se
   valida por diseño: se procesa un archivo = un documento).
4. INE: debe traer 2 páginas en el mismo PDF (frente y reverso) y la
   vigencia impresa no debe haber pasado ya.
5. CSF: debe traer 2 páginas en el mismo PDF y la fecha de emisión no debe
   tener más de 3 meses. Se lee también el "Estatus en el padrón" que la
   propia Constancia imprime (Activo / Suspendido / Cancelado) como
   verificación de autenticidad de primer nivel.
6. Comprobante de domicilio: no debe tener más de 3 meses de antigüedad.
7. Cuenta bancaria: se identifica el banco a partir de la CLABE (los
   primeros 3 dígitos son el código de institución bancaria, es más
   confiable que buscar el nombre del banco en el texto). Se valida que
   existan cuenta, CLABE y nombre del colaborador, y se rechaza si el banco
   es Nu, Spin by OXXO o Mercado Pago.

LIMITACIONES IMPORTANTES (léelas antes de confiar 100% en el resultado)
------------------------------------------------------------------------
- La verificación de que la carátula bancaria "incluya el logo del banco"
  es una revisión VISUAL; el OCR no puede confirmar que hay un logotipo.
  El script sí valida el banco (vía CLABE), cuenta, CLABE y nombre, pero
  marca el logo como "revisar visualmente".
- La autenticidad plena de la CSF ante el SAT requiere consultar el
  portal oficial (siat.sat.gob.mx); este script no hace esa consulta en
  vivo (no es un servicio público disponible para automatizar). Usa el
  estatus impreso en el documento como primera señal y genera, si el QR
  es legible, el enlace para que alguien lo confirme con un clic.
- El OCR de documentos escaneados no es perfecto. Toda fecha o nombre
  extraído por OCR debe leerse como "propuesta a confirmar", no como
  verdad absoluta — por eso el reporte siempre incluye el fragmento de
  texto de donde salió el dato.

USO
---
    python3 validador_expediente.py --nombre "Alam Naresh Poot Cauich" \
        --rfc POCA0201084WA --curp POCA020108HYNTCLA6 \
        --salida expediente_alam.xlsx \
        archivo1.pdf archivo2.pdf ...

Requiere: pdfplumber, pytesseract, pdf2image, openpyxl, y tener instalados
en el sistema los binarios `tesseract` (con el paquete de idioma `spa`) y
`poppler-utils` (pdftoppm). En Debian/Ubuntu:
    apt-get install -y tesseract-ocr tesseract-ocr-spa poppler-utils
    pip install pdfplumber pytesseract pdf2image openpyxl
"""

import argparse
import datetime
import os
import re
import sys
import unicodedata

import pdfplumber
from pdf2image import convert_from_path
from PIL import ImageOps, ImageStat, ImageFilter
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HOY = datetime.date.today()

# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------

def normaliza(txt):
    """Mayúsculas, sin acentos, espacios colapsados."""
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", txt).upper().strip()


MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

MESES_ABREV = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}


def busca_fechas(texto):
    """Devuelve una lista de (date, fragmento_texto) encontradas en el texto,
    soportando '20/05/2026' y '20 de mayo de 2026' / '20 DE MAYO DE 2026'."""
    fechas = []
    t = texto or ""

    for m in re.finditer(r"(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})", t):
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            fechas.append((datetime.date(y, mo, d), m.group(0)))
        except ValueError:
            pass

    patron_letras = re.compile(
        r"(\d{1,2})\s*(?:DE|DEL)?\s*([A-ZÑ]+)\s*(?:DE|DEL)?\s*(\d{4})",
        re.IGNORECASE,
    )
    for m in patron_letras.finditer(normaliza(t)):
        dia, mes_txt, anio = m.group(1), m.group(2), m.group(3)
        mes = MESES.get(mes_txt.upper())
        if mes:
            try:
                fechas.append((datetime.date(int(anio), mes, int(dia)), m.group(0)))
            except ValueError:
                pass

    # formato abreviado tipo recibo (ej. CFE): "03 MAY 26" o "03/MAY/26"
    patron_abrev = re.compile(r"(\d{1,2})\s*[/\s]\s*([A-Z]{3})\s*[/\s]\s*(\d{2})\b")
    for m in patron_abrev.finditer(normaliza(t)):
        dia, mes_txt, anio2 = m.group(1), m.group(2), m.group(3)
        mes = MESES_ABREV.get(mes_txt.upper())
        if mes:
            anio = 2000 + int(anio2)
            try:
                fechas.append((datetime.date(anio, mes, int(dia)), m.group(0)))
            except ValueError:
                pass

    return fechas


def fecha_mas_reciente_razonable(fechas, no_futuras=True):
    """De una lista de (date, frag), regresa la más reciente que no sea del
    futuro lejano (para evitar folios/números mal interpretados como fecha)."""
    candidatas = [f for f in fechas if f[0].year >= 2015 and f[0].year <= HOY.year + 1]
    if no_futuras:
        candidatas = [f for f in candidatas if f[0] <= HOY]
    if not candidatas:
        return None
    return max(candidatas, key=lambda f: f[0])


# ---------------------------------------------------------------------------
# Extracción de texto (nativo + OCR de respaldo)
# ---------------------------------------------------------------------------

# Límite de tiempo (segundos) por cada llamada a Tesseract. Sin esto, una
# imagen "difícil" (foto muy ruidosa, documento con un layout raro) puede
# hacer que el análisis de la página se tarde muchísimo y — como ya pasó una
# vez con Render — deje la generación del Excel como pasmada. Con el límite,
# si una sola llamada se pasa de tiempo simplemente se descarta ese intento
# (se trata como si no hubiera podido leer nada) y se sigue con el siguiente,
# en vez de quedarse trabada ahí.
OCR_TIMEOUT_SEGUNDOS = 12


def _ocr_con_confianza(imagen, lang="spa", config=""):
    """Corre OCR y regresa (texto, confianza_promedio_0_a_100).

    A diferencia de adivinar qué tan "limpio" se ve un texto contando tipos
    de caracteres (eso falla: un OCR que lee mal puede seguir escupiendo
    letras y espacios "normales", solo que las palabras equivocadas — se
    ve limpio pero está mal), aquí se usa la confianza que el propio
    Tesseract calcula por palabra (0-100, viene de image_to_data) y se
    promedia. Es la señal más confiable para decidir si vale la pena
    escalar a un pase más pesado, o para elegir cuál de varias
    configuraciones de Tesseract leyó mejor una misma imagen."""
    try:
        datos = pytesseract.image_to_data(
            imagen, lang=lang, config=config, timeout=OCR_TIMEOUT_SEGUNDOS,
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return "", -1.0

    lineas = {}
    confianzas = []
    n = len(datos.get("text", []))
    for i in range(n):
        texto = (datos["text"][i] or "").strip()
        if not texto:
            continue
        clave_linea = (datos["block_num"][i], datos["par_num"][i], datos["line_num"][i])
        lineas.setdefault(clave_linea, []).append(texto)
        try:
            conf = float(datos["conf"][i])
        except (TypeError, ValueError):
            conf = -1.0
        if conf >= 0:
            confianzas.append(conf)

    texto_completo = "\n".join(" ".join(palabras) for palabras in lineas.values())
    confianza_prom = sum(confianzas) / len(confianzas) if confianzas else -1.0
    return texto_completo, confianza_prom


def _preprocesa_para_ocr(imagen, nivel="ligero"):
    """Prepara una imagen antes de mandarla al OCR.
    - 'ligero': escala de grises + autocontraste. No es OCR, es solo
      procesamiento de imagen (barato en CPU), así que se aplica siempre
      desde el pase rápido — ayuda bastante con fotos de celular con poco
      contraste o fondos de color, sin costo extra relevante.
    - 'fuerte': lo anterior + binarización (blanco/negro, usando el brillo
      promedio de la imagen para fijar el umbral) + nitidez. Es lo que más
      ayuda con documentos que a simple vista se ven bien pero el OCR lee
      mal: fondos de color en credenciales, sombras, brillo del flash, papel
      térmico desteñido. Es más pesado, por eso solo se usa en el pase HD, y
      solo para las páginas que ya fallaron el pase rápido."""
    gris = ImageOps.grayscale(imagen)
    # OJO: cutoff>0 aquí puede ser CONTRAPRODUCENTE. Se probó con cutoff=2
    # (recorta el 2% más oscuro/claro del histograma antes de estirar el
    # contraste) y en un documento de fondo casi blanco con poco texto
    # oscuro, ese recorte metía un patrón de moteado en las letras que hacía
    # que Tesseract no reconociera NADA en un documento perfectamente
    # legible a simple vista. cutoff=0 (usar el mínimo/máximo real de la
    # imagen para estirar el contraste, sin recortar nada) no tiene ese
    # problema y sigue ayudando con fotos de bajo contraste.
    gris = ImageOps.autocontrast(gris, cutoff=0)
    if nivel == "ligero":
        return gris
    brillo_medio = ImageStat.Stat(gris).mean[0]
    umbral = max(100, min(200, brillo_medio * 0.85))
    binaria = gris.point(lambda x: 255 if x > umbral else 0)
    return binaria.filter(ImageFilter.SHARPEN)


def _corrige_rotacion(imagen):
    """Detecta si la imagen viene rotada 90/180/270° —común cuando el
    candidato fotografía su credencial o comprobante con el celular en la
    orientación equivocada, lo cual hace que Tesseract no reconozca casi nada
    aunque el documento sea perfectamente legible— y la corrige antes del
    OCR. Si Tesseract no logra determinar la orientación (imagen muy
    ruidosa) simplemente se deja igual."""
    try:
        osd = pytesseract.image_to_osd(
            imagen, output_type=pytesseract.Output.DICT, timeout=OCR_TIMEOUT_SEGUNDOS
        )
        angulo = int(osd.get("rotate", 0) or 0)
        # PIL's Image.rotate(x) gira en sentido antihorario x grados; el
        # campo "rotate" que regresa Tesseract ya viene en la convención que
        # hace falta pasarle directo (probado empíricamente: usar -angulo
        # deja la imagen al revés). Por eso aquí es rotate(angulo), sin signo
        # invertido.
        if angulo:
            return imagen.rotate(angulo, expand=True)
    except Exception:
        pass
    return imagen


def extraer_texto_pdf(ruta):
    """Regresa (lista_de_texto_por_pagina, num_paginas, uso_ocr_por_pagina).

    Estrategia de OCR en 2 pasos:
      1) Pase rápido (200 dpi + gris/autocontraste ligero) — resuelve la
         gran mayoría de documentos escaneados o fotografiados con luz
         decente, y sigue siendo barato en CPU (funciona incluso en un
         hosting con muy pocos recursos).
      2) Solo para las páginas que el pase rápido no leyó bien —ya sea
         porque salió muy poco texto, o porque Tesseract mismo reporta poca
         confianza en lo que leyó (típico de una foto con glare, sombra o
         fondo de color, aunque el documento sea legible a simple vista)—
         un pase HD (300 dpi) que además:
           - corrige automáticamente la rotación (fotos tomadas de lado),
           - prueba la imagen con y sin binarizar, y
           - prueba dos configuraciones de segmentación de Tesseract
             (--psm 3 y 6),
         quedándose con la combinación de mayor confianza reportada por el
         propio Tesseract (no la primera que salga ni la más larga). Cada
         intento individual tiene un límite de tiempo (OCR_TIMEOUT_SEGUNDOS)
         para que una imagen realmente mala no trabe la generación completa
         del Excel.
    """
    paginas_texto = []
    ocr_usado = []
    with pdfplumber.open(ruta) as pdf:
        num_paginas = len(pdf.pages)
        for pagina in pdf.pages:
            texto = (pagina.extract_text() or "").strip()
            paginas_texto.append(texto)
            ocr_usado.append(False)

    # Umbral de confianza (0-100, escala propia de Tesseract) bajo el cual se
    # considera que un pase de OCR "no se puede confiar" y conviene escalar
    # al siguiente pase, aunque ya haya salido algo de texto.
    CONFIANZA_MINIMA = 60

    necesita_ocr = [i for i, t in enumerate(paginas_texto) if len(t) < 20]
    if necesita_ocr:
        try:
            imagenes_rapidas = convert_from_path(ruta, dpi=200)
        except Exception as e:
            imagenes_rapidas = []
            print(f"  [aviso] no se pudo rasterizar para OCR ({e})", file=sys.stderr)

        reintentar = []
        for i in necesita_ocr:
            if i >= len(imagenes_rapidas):
                continue
            imagen_prep = _preprocesa_para_ocr(imagenes_rapidas[i], nivel="ligero")
            texto_ocr, confianza = _ocr_con_confianza(imagen_prep, lang="spa")
            if len(texto_ocr) > len(paginas_texto[i]):
                paginas_texto[i] = texto_ocr
                ocr_usado[i] = True
            # Se reintenta con el pase HD no solo si vino muy poco texto,
            # sino también si Tesseract mismo reporta poca confianza en lo
            # que leyó —eso es justo lo que pasa con documentos legibles a
            # simple vista pero que el pase rápido lee mal (poca luz, glare,
            # fondo de color)—.
            if len(texto_ocr) < 20 or confianza < CONFIANZA_MINIMA:
                reintentar.append(i)

        if reintentar:
            try:
                imagenes_hd = convert_from_path(ruta, dpi=300)
            except Exception as e:
                imagenes_hd = []
                print(f"  [aviso] no se pudo rasterizar en HD para OCR ({e})", file=sys.stderr)
            for i in reintentar:
                if i >= len(imagenes_hd):
                    continue
                imagen_hd = _corrige_rotacion(imagenes_hd[i])
                # Se prueban dos variantes de preprocesamiento (con y sin
                # binarizar — binarizar ayuda mucho con fondos de color pero
                # en documentos de fondo claro a veces borra más de lo que
                # ayuda) combinadas con varias configuraciones de
                # segmentación de página de Tesseract, y se elige la que dé
                # mayor confianza — no la primera que salga ni la más larga.
                variantes_imagen = [
                    _preprocesa_para_ocr(imagen_hd, nivel="ligero"),
                    _preprocesa_para_ocr(imagen_hd, nivel="fuerte"),
                ]
                mejor_texto, mejor_confianza = paginas_texto[i], -1.0
                mejoro = False
                terminar = False
                for imagen_candidata in variantes_imagen:
                    for psm in ("3", "6"):
                        try:
                            candidato, confianza = _ocr_con_confianza(
                                imagen_candidata, lang="spa", config=f"--psm {psm}"
                            )
                        except Exception as e:
                            candidato, confianza = "", -1.0
                            print(f"  [aviso] OCR (HD, psm {psm}) falló en página {i+1} ({e})", file=sys.stderr)
                        if candidato and len(candidato) >= 20 and confianza > mejor_confianza:
                            mejor_texto, mejor_confianza = candidato, confianza
                            mejoro = True
                        # Ya se ve confiable y con longitud razonable: no
                        # vale la pena seguir probando más configuraciones.
                        if mejor_confianza >= 75 and len(mejor_texto) >= 40:
                            terminar = True
                            break
                    if terminar:
                        break
                if mejoro:
                    paginas_texto[i] = mejor_texto
                    ocr_usado[i] = True

    return paginas_texto, num_paginas, ocr_usado


# ---------------------------------------------------------------------------
# Clasificación de documento
# ---------------------------------------------------------------------------

# clave -> (nombre legible, lista de palabras/frases clave, es obligatorio)
CATEGORIAS = {
    "CV": ("CV", ["EXPERIENCIA LABORAL", "REFERENCIAS LABORALES", "PERFIL PROFESIONAL", "CURRICULUM"], True),
    "ACTA_NACIMIENTO": ("Acta de nacimiento", ["ACTA DE NACIMIENTO", "REGISTRO CIVIL", "OFICIALIA", "NACIMIENTOS"], True),
    "INE": ("INE", ["INSTITUTO NACIONAL ELECTORAL", "CREDENCIAL PARA VOTAR", "CLAVE DE ELECTOR"], True),
    "COMPROBANTE_DOMICILIO": ("Comprobante de domicilio", ["COMISION FEDERAL DE ELECTRICIDAD", "CFE", "TELMEX", "RECIBO", "TOTAL A PAGAR", "PERIODO FACTURADO", "IZZI", "TELEFONOS DE MEXICO", "AGUA"], True),
    "COMPROBANTE_ESTUDIOS": ("Comprobante de estudios", ["CERTIFICADO DE ESTUDIOS", "CURSO Y ACREDITO", "UNIVERSIDAD", "LICENCIATURA", "CEDULA PROFESIONAL", "SECRETARIA DE EDUCACION", "BACHILLERATO", "PROMEDIO GENERAL"], True),
    "CURP": ("CURP", ["CLAVE UNICA DE REGISTRO DE POBLACION", "CURP CERTIFICADA", "CLAVE:"], True),
    "CSF": ("CSF", ["CONSTANCIA DE SITUACION FISCAL", "CEDULA DE IDENTIFICACION FISCAL", "REGISTRO FEDERAL DE CONTRIBUYENTES"], True),
    "NSS": ("NSS", ["NUMERO DE SEGURIDAD SOCIAL", "INSTITUTO MEXICANO DEL SEGURO SOCIAL", "IMSS"], True),
    "CUENTA_BANCARIA": ("Cuenta bancaria", ["CLABE", "ESTADO DE CUENTA", "NO. DE CUENTA", "CARATULA"], True),
    "INFONAVIT_FONACOT": ("Aviso de retención Infonavit/Fonacot", ["INFONAVIT", "FONACOT", "AVISO DE RETENCION"], False),
    # condicionales: aplican según el puesto (entrenador, barbero, estilista) o la situación del candidato
    "CERTIFICADO_MEDICO": ("Certificado médico", ["CERTIFICADO MEDICO", "RECONOCIMIENTO MEDICO", "MEDICO CIRUJANO"], False),
    "CERTIFICADO_INSTRUCTOR": ("Certificado de entrenador / barbero / estilista", ["CERTIFICADO", "FITNESS COACH", "ENTRENADOR", "ESTILISTA", "BARBERO", "BARBER", "COSMETOLOGIA", "COSMETOLOGO", "DIPLOMADO"], False),
    "CONSTANCIA_LABORAL": ("Constancia(s) laboral(es) / cartas de referencia", ["CONSTANCIA LABORAL", "CARTA LABORAL", "HACE CONSTAR QUE", "CARTA DE RECOMENDACION", "RECURSOS HUMANOS"], False),
}

# orden e integrantes de la checklist obligatoria que pidió el negocio
CHECKLIST_OBLIGATORIO = [
    "CV", "ACTA_NACIMIENTO", "INE", "COMPROBANTE_DOMICILIO",
    "COMPROBANTE_ESTUDIOS", "CURP", "CSF", "NSS", "CUENTA_BANCARIA",
    "INFONAVIT_FONACOT", "CERTIFICADO_MEDICO", "CERTIFICADO_INSTRUCTOR",
    "CONSTANCIA_LABORAL",
]

# de estos, cuáles son condicionales (no siempre aplican) en vez de siempre-obligatorios
CONDICIONALES = {"INFONAVIT_FONACOT", "CERTIFICADO_MEDICO", "CERTIFICADO_INSTRUCTOR", "CONSTANCIA_LABORAL"}


def clasificar(texto_completo, nombre_archivo):
    t = normaliza(texto_completo)
    nombre_arch_norm = normaliza(nombre_archivo)
    mejor_clave, mejor_score = "DESCONOCIDO", 0
    for clave, (_, palabras, _) in CATEGORIAS.items():
        score = sum(1 for p in palabras if normaliza(p) in t)
        # pequeño empujón si el nombre del archivo ya lo sugiere
        pistas_nombre = {
            "CV": ["CV", "CURRICULUM"], "ACTA_NACIMIENTO": ["ACTA"], "INE": ["INE", "IFE"],
            "COMPROBANTE_DOMICILIO": ["DOMICILIO", "RECIBO", "CFE", "LOCALIZACION"],
            "COMPROBANTE_ESTUDIOS": ["GRADO", "ESTUDIOS", "TITULO", "CEDULA"],
            "CURP": ["CURP"], "CSF": ["CSF", "CONSTANCIA", "FISCAL"], "NSS": ["NSS", "SEGURIDAD SOCIAL", "LOCALIZACION"],
            "CUENTA_BANCARIA": ["CUENTA", "BANCO", "ESTADO DE CUENTA", "CARATULA"],
            "INFONAVIT_FONACOT": ["INFONAVIT", "FONACOT"],
            "CERTIFICADO_INSTRUCTOR": ["CERTIFICADO", "BARBER", "ESTILISTA", "COACH"],
            "CERTIFICADO_MEDICO": ["MEDICO"],
            "CONSTANCIA_LABORAL": ["LABORAL", "CONSTANCIA"],
        }
        for pista in pistas_nombre.get(clave, []):
            if pista in nombre_arch_norm:
                score += 1
        if score > mejor_score:
            mejor_clave, mejor_score = clave, score
    return mejor_clave if mejor_score > 0 else "DESCONOCIDO"


# ---------------------------------------------------------------------------
# Nombre: extracción y comparación
# ---------------------------------------------------------------------------

def tokens_nombre(nombre):
    return [t for t in normaliza(nombre).split(" ") if len(t) > 1]


def nombre_coincide(texto_doc, nombre_candidato):
    """Compara por tokens: cuenta cuántas palabras del nombre del candidato
    aparecen en el texto del documento. Tolerante a orden y a acentos."""
    tks = tokens_nombre(nombre_candidato)
    if not tks:
        return None, 0
    texto_norm = normaliza(texto_doc)
    encontrados = sum(1 for tk in tks if re.search(r"\b" + re.escape(tk) + r"\b", texto_norm))
    proporcion = encontrados / len(tks)
    if proporcion >= 0.75:
        return True, proporcion
    elif proporcion >= 0.4:
        return None, proporcion  # dudoso -> revisar a mano
    else:
        return False, proporcion


# ---------------------------------------------------------------------------
# Reglas específicas por documento
# ---------------------------------------------------------------------------

CODIGOS_CLABE_BANCOS = {
    "002": "Banamex/Citibanamex", "006": "Bancomext", "009": "Banobras",
    "012": "BBVA México", "014": "Santander", "019": "Banjercito",
    "021": "HSBC", "030": "Bajío", "036": "Inbursa", "037": "Interbanco",
    "042": "Mifel", "044": "Scotiabank", "058": "Banregio", "059": "Invex",
    "060": "Bansi", "062": "Afirme", "072": "Banorte", "102": "Multiva",
    "103": "American Express", "106": "Bank of America", "108": "MUFG",
    "110": "JP Morgan", "112": "BMONEX", "113": "VE POR MAS", "124": "Deutsche Bank",
    "127": "Banco Azteca", "128": "Autofin", "129": "Barclays", "130": "Compartamos",
    "131": "Banco Famsa", "132": "BMULTIVA", "133": "Actinver", "134": "Wal-Mart (Bancoppel línea)",
    "135": "Nafin", "136": "Intercam Banco", "137": "Bankaool", "138": "ABC Capital",
    "140": "Consubanco", "141": "Volkswagen Bank", "143": "CIBanco", "145": "BBASE",
    "147": "Bankaool", "148": "PagaTodo", "150": "Inmobiliario", "151": "Donde",
    "152": "Bancrea", "154": "Banco Covalto", "155": "ICBC", "156": "Sabadell",
    "157": "Shinhan", "158": "Mizuho Bank", "160": "Banco S3", "166": "Bansefi/Bienestar",
    "168": "Hipotecaria Federal",
    # Fintechs / billeteras que el negocio no acepta como cuenta de nómina
    "638": "NVIO Pagos México (Nu)", "722": "Mercado Pago W Digital", "728": "Spin by OXXO",
    "659": "Openpay/Klar (según convenio)", "646": "STP (posible operador de una fintech)",
}

BANCOS_EXCLUIDOS = {"NVIO Pagos México (Nu)", "Mercado Pago W Digital", "Spin by OXXO"}
NOMBRES_EXCLUIDOS_TEXTO = ["NU MEXICO", "NU BANK", "SPIN BY OXXO", "MERCADO PAGO", "MERCADOPAGO"]


def analiza_cuenta_bancaria(texto_completo, nombre_candidato):
    obs = []
    t_norm = normaliza(texto_completo)

    m_clabe = re.search(r"CLABE[^0-9]{0,15}(\d[\d\s]{16,22}\d)", t_norm)
    clabe = re.sub(r"\s", "", m_clabe.group(1)) if m_clabe else None
    if clabe and len(clabe) >= 18:
        clabe = clabe[:18]

    banco = None
    if clabe:
        banco = CODIGOS_CLABE_BANCOS.get(clabe[:3], f"Código de banco no identificado ({clabe[:3]})")

    m_cuenta = re.search(r"(?:NO\.?\s*DE\s*CUENTA|NUMERO\s*DE\s*CUENTA|CUENTA)\s*:?\s*(\d{6,20})", t_norm)
    numero_cuenta = m_cuenta.group(1) if m_cuenta else None
    tiene_cuenta = numero_cuenta is not None
    nombre_ok, proporcion = nombre_coincide(texto_completo, nombre_candidato)

    banco_rechazado = False
    if banco in BANCOS_EXCLUIDOS:
        banco_rechazado = True
    elif not clabe:
        # sin CLABE detectable por regex, buscamos mención directa de banco excluido en el encabezado
        encabezado = t_norm[:600]
        if any(k in encabezado for k in NOMBRES_EXCLUIDOS_TEXTO):
            banco_rechazado = True
            banco = "Nu / Spin / Mercado Pago (detectado por texto, sin CLABE confirmada)"

    if banco_rechazado:
        obs.append(f"Banco NO permitido para depósito de nómina: {banco}.")
    elif banco is None:
        obs.append("No se detectó CLABE ni banco; revisar manualmente que sea un banco permitido.")
    else:
        obs.append(f"Banco detectado: {banco} (permitido).")

    if not clabe:
        obs.append("No se encontró una CLABE de 18 dígitos legible.")
    if not tiene_cuenta:
        obs.append("No se encontró un número de cuenta explícito.")
    obs.append("El logo del banco no puede confirmarse por OCR: revisar visualmente el PDF.")

    return {
        "banco": banco,
        "banco_rechazado": banco_rechazado,
        "clabe_detectada": clabe,
        "numero_cuenta": numero_cuenta,
        "tiene_cuenta": tiene_cuenta,
        "nombre_coincide": nombre_ok,
        "observaciones": " ".join(obs),
    }


def analiza_csf(paginas_texto):
    texto_completo = "\n".join(paginas_texto)
    t_norm = normaliza(texto_completo)
    obs = []

    # el PDF de la CSF a veces pierde los espacios entre palabras al extraer texto
    # ("Estatusenelpadrón:ACTIVO"), así que probamos con y sin espacios.
    m_estatus = re.search(r"ESTATUS\s*EN\s*EL\s*PADRON\s*:?\s*([A-Z]+)", t_norm)
    estatus = m_estatus.group(1) if m_estatus else None
    if estatus == "ACTIVO":
        obs.append("La CSF indica estatus ACTIVO en el padrón del SAT.")
    elif estatus:
        obs.append(f"La CSF indica estatus '{estatus}' (revisar, no es ACTIVO).")
    else:
        obs.append("No se pudo leer el estatus del contribuyente en el texto; revisar manualmente.")

    fechas = busca_fechas(texto_completo)
    emision = fecha_mas_reciente_razonable(fechas)
    dentro_3_meses = None
    if emision:
        dias = (HOY - emision[0]).days
        dentro_3_meses = dias <= 92
        obs.append(f"Fecha de emisión detectada: {emision[0].isoformat()} ({dias} días de antigüedad).")
    else:
        obs.append("No se detectó una fecha de emisión clara; revisar manualmente.")

    obs.append("Autenticidad ante el SAT: no se consulta en vivo el portal del SAT desde este script; "
                "si el documento trae QR legible, ábranlo en el validador oficial para confirmar.")

    return {
        "estatus": estatus,
        "activo": estatus == "ACTIVO",
        "fecha_emision": emision[0] if emision else None,
        "dentro_3_meses": dentro_3_meses,
        "num_paginas_ok": len(paginas_texto) >= 2,
        "observaciones": " ".join(obs),
    }


def analiza_ine(paginas_texto):
    obs = []

    dos_paginas = len(paginas_texto) >= 2
    if dos_paginas:
        obs.append("PDF trae 2 páginas: se asume frente y reverso en un solo archivo.")
    else:
        obs.append("El PDF trae solo 1 página: falta el reverso (o viene en archivo separado).")

    # La vigencia SIEMPRE se busca en la cara frontal (página 1) — el reverso
    # es la franja MRZ y no trae el campo "VIGENCIA", así que buscarla ahí
    # solo metería ruido. Si el PDF llegó con una sola página, se analiza esa
    # misma como frente.
    texto_frente = paginas_texto[0] if paginas_texto else ""
    t_norm = normaliza(texto_frente)

    # El OCR de la credencial suele meter ruido entre "VIGENCIA" y los años
    # (números de sección, caracteres mal leídos), así que buscamos los dos
    # años (19xx/20xx) más cercanos después de la palabra VIGENCIA en vez de
    # exigir que estén pegados a ella.
    vigente = None
    anio_fin = None
    idx = t_norm.find("VIGENCIA")
    if idx != -1:
        ventana = t_norm[idx: idx + 80]
        anios = re.findall(r"\b(19\d{2}|20\d{2})\b", ventana)
        if len(anios) >= 2:
            anio_inicio, anio_fin = int(anios[-2]), int(anios[-1])
            vigente = anio_fin >= HOY.year
            dias_para_vencer = (datetime.date(anio_fin, 12, 31) - HOY).days
            if vigente:
                obs.append(f"Vigencia impresa en la cara frontal: {anio_inicio}-{anio_fin} (vigente hoy {HOY.isoformat()}).")
            else:
                obs.append(f"Vigencia impresa en la cara frontal: {anio_inicio}-{anio_fin} — VENCIDA (venció hace {abs(dias_para_vencer)} días respecto a hoy {HOY.isoformat()}).")
        else:
            obs.append("Se encontró la palabra VIGENCIA en la cara frontal pero no dos años legibles junto a ella; revisar manualmente.")
    else:
        obs.append("No se detectó el campo de vigencia en la cara frontal; revisar manualmente.")

    return {
        "vigente": vigente,
        "vigencia_anio_fin": anio_fin,
        "dos_paginas": dos_paginas,
        "observaciones": " ".join(obs),
    }


def analiza_fecha_limite(texto_completo, dias_limite, etiqueta):
    """Busca el campo de vigencia/fecha en el texto y lo compara contra el día
    de hoy. Regresa (dentro_del_limite, texto_de_observacion, fecha_encontrada)."""
    fechas = busca_fechas(texto_completo)
    fecha = fecha_mas_reciente_razonable(fechas)
    if not fecha:
        return None, f"No se detectó el campo de vigencia/fecha en {etiqueta}; revisar manualmente.", None
    dias = (HOY - fecha[0]).days
    dentro = dias <= dias_limite
    if dentro:
        obs = (f"Vigencia de {etiqueta}: fecha detectada {fecha[0].isoformat()}, hoy es {HOY.isoformat()} "
               f"({dias} días de antigüedad, dentro del límite de {dias_limite} días).")
    else:
        obs = (f"Vigencia de {etiqueta}: fecha detectada {fecha[0].isoformat()}, hoy es {HOY.isoformat()} "
               f"— FUERA DE VIGENCIA ({dias} días de antigüedad, supera el límite de {dias_limite} días).")
    return dentro, obs, fecha[0]


# ---------------------------------------------------------------------------
# Procesamiento principal
# ---------------------------------------------------------------------------

def procesar_documento(ruta, nombre_candidato):
    nombre_archivo = os.path.basename(ruta)
    paginas_texto, num_paginas, ocr_usado = extraer_texto_pdf(ruta)
    texto_completo = "\n".join(paginas_texto)
    legible = any(len(p.strip()) >= 20 for p in paginas_texto)
    clave = clasificar(texto_completo, nombre_archivo)
    nombre_legible, _, _ = CATEGORIAS.get(clave, ("Desconocido / no identificado", [], False))

    fila = {
        "archivo": nombre_archivo,
        "categoria_clave": clave,
        "categoria": nombre_legible,
        "num_paginas": num_paginas,
        "uso_ocr": any(ocr_usado),
        "legible": legible,
        "texto_muestra": texto_completo[:400].replace("\n", " ").strip(),
        "nombre_coincide": None,
        "detalle": "",
    }

    if not legible:
        fila["detalle"] = "El documento no pudo leerse (ni texto nativo ni OCR); solicitar de nuevo, escaneo/foto más nítida."
        return fila

    excluye_nombre = clave == "COMPROBANTE_DOMICILIO"
    if not excluye_nombre:
        coincide, proporcion = nombre_coincide(texto_completo, nombre_candidato)
        fila["nombre_coincide"] = coincide
        fila["detalle"] += f"Coincidencia de nombre: {proporcion*100:.0f}% de las palabras del nombre capturado se encontraron en el documento. "

    detalles_extra = []

    if clave == "CSF":
        r = analiza_csf(paginas_texto)
        fila["num_paginas_ok"] = r["num_paginas_ok"]
        fila["vigencia_ok"] = r["dentro_3_meses"]
        fila["vigencia_fecha_texto"] = r["fecha_emision"].isoformat() if r["fecha_emision"] else None
        fila["estatus_sat"] = r["estatus"]
        detalles_extra.append(r["observaciones"])
        if not r["num_paginas_ok"]:
            detalles_extra.append("Falta la segunda hoja de la CSF en este PDF (debe traer ambas en 1 solo archivo).")

    elif clave == "INE":
        # La vigencia SIEMPRE se revisa en la cara frontal (página 1) contra
        # la fecha de hoy; que el PDF traiga ambas caras solo se checa por
        # separado (num_paginas_ok) para no mezclar los dos criterios.
        r = analiza_ine(paginas_texto)
        fila["vigencia_ok"] = r["vigente"]
        fila["vigencia_fecha_texto"] = f"Vigente hasta {r['vigencia_anio_fin']}" if r["vigencia_anio_fin"] else None
        fila["num_paginas_ok"] = r["dos_paginas"]
        detalles_extra.append(r["observaciones"])

    elif clave == "COMPROBANTE_DOMICILIO":
        dentro, obs, fecha = analiza_fecha_limite(texto_completo, 92, "el comprobante de domicilio")
        fila["vigencia_ok"] = dentro
        fila["vigencia_fecha_texto"] = fecha.isoformat() if fecha else None
        detalles_extra.append(obs)

    elif clave == "CUENTA_BANCARIA":
        r = analiza_cuenta_bancaria(texto_completo, nombre_candidato)
        fila["banco"] = r["banco"]
        fila["banco_rechazado"] = r["banco_rechazado"]
        fila["clabe"] = r["clabe_detectada"]
        fila["numero_cuenta"] = r["numero_cuenta"]
        detalles_extra.append(r["observaciones"])
        if num_paginas > 2:
            detalles_extra.append(
                f"El archivo trae {num_paginas} páginas (estado de cuenta completo). "
                "Se recomienda pedir solo la carátula (1 página) para no exponer el detalle de movimientos."
            )

    elif clave == "ACTA_NACIMIENTO":
        dentro, obs, fecha = analiza_fecha_limite(texto_completo, 365 * 5, "el acta de nacimiento")
        fila["vigencia_ok"] = dentro
        fila["vigencia_fecha_texto"] = fecha.isoformat() if fecha else None
        detalles_extra.append(obs + " (regla de 5 años detectada como práctica común; confirmar si aplica formalmente).")

    fila["detalle"] += " ".join(detalles_extra)
    return fila


def construir_reporte(candidato, filas):
    encontrados_por_categoria = {}
    for fila in filas:
        encontrados_por_categoria.setdefault(fila["categoria_clave"], []).append(fila)

    checklist = []
    for clave in CHECKLIST_OBLIGATORIO:
        nombre_legible, _, obligatorio = CATEGORIAS[clave]
        docs = encontrados_por_categoria.get(clave, [])
        checklist.append({
            "clave": clave,
            "categoria": nombre_legible,
            "obligatorio": obligatorio,
            "recibido": len(docs) > 0,
            "archivos": [d["archivo"] for d in docs],
        })

    extra = []
    for clave, docs in encontrados_por_categoria.items():
        if clave not in CHECKLIST_OBLIGATORIO:
            nombre_legible, _, _ = CATEGORIAS.get(clave, ("Desconocido / no identificado", [], False))
            extra.append({"categoria": nombre_legible, "archivos": [d["archivo"] for d in docs]})

    return checklist, extra


# ---------------------------------------------------------------------------
# Salida a Excel
# ---------------------------------------------------------------------------

VERDE = PatternFill("solid", fgColor="C6EFCE")
ROJO = PatternFill("solid", fgColor="FFC7CE")
AMARILLO = PatternFill("solid", fgColor="FFEB9C")
GRIS_ENCABEZADO = PatternFill("solid", fgColor="1F6B4C")
FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)


def _set_encabezados(ws, encabezados, fila=1):
    for col, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = FUENTE_ENCABEZADO
        c.fill = GRIS_ENCABEZADO
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[fila].height = 30


def _autoancho(ws, anchos):
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def generar_excel(candidato, filas, checklist, extra, ruta_salida):
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Expediente de reclutamiento — validación automática"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Candidato: {candidato['nombre']}"
    ws["A3"] = f"RFC capturado: {candidato.get('rfc') or '—'}    CURP capturado: {candidato.get('curp') or '—'}"
    ws["A4"] = f"Generado: {HOY.isoformat()}"

    faltantes = [c for c in checklist if c["obligatorio"] and not c["recibido"]]
    completo = len(faltantes) == 0
    ws["A6"] = "Estado de completitud:"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = "COMPLETO" if completo else f"INCOMPLETO — faltan {len(faltantes)} documento(s)"
    ws["B6"].fill = VERDE if completo else ROJO
    ws["B6"].font = Font(bold=True)

    _set_encabezados(ws, ["Documento", "Obligatorio", "¿Recibido?", "Archivo(s)"], fila=8)
    r = 9
    for item in checklist:
        ws.cell(row=r, column=1, value=item["categoria"])
        ws.cell(row=r, column=2, value="Sí" if item["obligatorio"] else "Condicional (solo si aplica)")
        celda_recibido = ws.cell(row=r, column=3, value="Sí" if item["recibido"] else "No")
        if item["obligatorio"]:
            celda_recibido.fill = VERDE if item["recibido"] else ROJO
        else:
            celda_recibido.fill = VERDE if item["recibido"] else AMARILLO
        ws.cell(row=r, column=4, value=", ".join(item["archivos"]) or "—")
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Documentos faltantes (obligatorios):").font = Font(bold=True)
    r += 1
    if faltantes:
        for item in faltantes:
            c = ws.cell(row=r, column=1, value="• " + item["categoria"])
            c.fill = ROJO
            r += 1
    else:
        ws.cell(row=r, column=1, value="Ninguno — la documentación obligatoria está completa.")
        r += 1

    condicionales_faltantes = [c for c in checklist if (not c["obligatorio"]) and not c["recibido"]]
    r += 1
    ws.cell(row=r, column=1, value="Documentos condicionales NO recibidos (confirmar con el candidato si le aplican):").font = Font(bold=True)
    r += 1
    if condicionales_faltantes:
        for item in condicionales_faltantes:
            c = ws.cell(row=r, column=1, value="• " + item["categoria"])
            c.fill = AMARILLO
            r += 1
    else:
        ws.cell(row=r, column=1, value="Ninguno pendiente.")
        r += 1

    if extra:
        r += 1
        ws.cell(row=r, column=1, value="Documentos adicionales recibidos (no están en la lista de arriba):").font = Font(bold=True, italic=True)
        r += 1
        for item in extra:
            ws.cell(row=r, column=1, value=item["categoria"])
            ws.cell(row=r, column=4, value=", ".join(item["archivos"]))
            r += 1

    # --- Vigencias verificadas (INE cara frontal, comprobante de domicilio,
    # CSF, acta) contra la fecha de hoy. Se listan aunque no se haya podido
    # leer la fecha, para que quede visible qué falta revisar a mano. -------
    DOCS_CON_REGLA_VIGENCIA = {"INE", "COMPROBANTE_DOMICILIO", "CSF", "ACTA_NACIMIENTO"}
    con_vigencia = [f for f in filas if f["categoria_clave"] in DOCS_CON_REGLA_VIGENCIA]
    r += 2
    ws.cell(row=r, column=1, value=f"Vigencias verificadas contra hoy ({HOY.isoformat()}):").font = Font(bold=True)
    r += 1
    if con_vigencia:
        _set_encabezados(ws, ["Documento", "Fecha / vigencia detectada", "Estado", "Archivo"], fila=r)
        r += 1
        for f in con_vigencia:
            ws.cell(row=r, column=1, value=f["categoria"])
            ws.cell(row=r, column=2, value=f.get("vigencia_fecha_texto") or "No se detectó")
            vig = f.get("vigencia_ok")
            c_estado = ws.cell(row=r, column=3, value="OK" if vig else "FUERA DE VIGENCIA — revisar" if vig is False else "Sin fecha detectada — revisar a mano")
            c_estado.fill = VERDE if vig else ROJO if vig is False else AMARILLO
            ws.cell(row=r, column=4, value=f["archivo"])
            r += 1
    else:
        ws.cell(row=r, column=1, value="No se recibieron documentos con regla de vigencia.")
        r += 1

    # --- Datos bancarios de la carátula, exportados como texto para que no
    # se corrompan en Excel (CLABE/cuenta con ceros a la izquierda, etc.) --
    bancarios = [f for f in filas if f["categoria_clave"] == "CUENTA_BANCARIA"]
    r += 2
    ws.cell(row=r, column=1, value="Datos bancarios (carátula) — CLABE y cuenta exportados como texto:").font = Font(bold=True)
    r += 1
    if bancarios:
        _set_encabezados(ws, ["Archivo", "Banco", "Número de cuenta", "CLABE"], fila=r)
        r += 1
        for f in bancarios:
            ws.cell(row=r, column=1, value=f["archivo"])
            ws.cell(row=r, column=2, value=f.get("banco") or "No identificado")
            c_cuenta = ws.cell(row=r, column=3, value=f.get("numero_cuenta") or "No detectado")
            c_cuenta.number_format = "@"  # forzar texto: evita notación científica o pérdida de ceros
            c_clabe = ws.cell(row=r, column=4, value=f.get("clabe") or "No detectada")
            c_clabe.number_format = "@"
            r += 1
    else:
        ws.cell(row=r, column=1, value="No se recibió carátula bancaria.")
        r += 1

    _autoancho(ws, [42, 26, 20, 45])

    ws2 = wb.create_sheet("Detalle por documento")
    encabezados2 = [
        "Archivo", "Documento identificado", "Páginas", "¿Usó OCR?", "Legible",
        "Nombre coincide", "Vigencia OK", "Fecha/vigencia detectada", "Banco (si aplica)",
        "Número de cuenta", "CLABE", "Observaciones",
    ]
    _set_encabezados(ws2, encabezados2)
    r = 2
    for fila in filas:
        ws2.cell(row=r, column=1, value=fila["archivo"])
        ws2.cell(row=r, column=2, value=fila["categoria"])
        ws2.cell(row=r, column=3, value=fila["num_paginas"])
        ws2.cell(row=r, column=4, value="Sí" if fila["uso_ocr"] else "No")

        c_leg = ws2.cell(row=r, column=5, value="Sí" if fila["legible"] else "NO — revisar")
        c_leg.fill = VERDE if fila["legible"] else ROJO

        nc = fila.get("nombre_coincide")
        if fila["categoria_clave"] == "COMPROBANTE_DOMICILIO":
            c_nom = ws2.cell(row=r, column=6, value="N/A (puede ser otra persona)")
            c_nom.fill = AMARILLO
        else:
            texto_nc = "Sí" if nc is True else "NO — revisar" if nc is False else "Dudoso — revisar"
            c_nom = ws2.cell(row=r, column=6, value=texto_nc)
            c_nom.fill = VERDE if nc is True else ROJO if nc is False else AMARILLO

        vig = fila.get("vigencia_ok")
        if vig is None:
            c_vig = ws2.cell(row=r, column=7, value="—")
        else:
            c_vig = ws2.cell(row=r, column=7, value="OK" if vig else "FUERA DE RANGO — revisar")
            c_vig.fill = VERDE if vig else ROJO

        ws2.cell(row=r, column=8, value=fila.get("vigencia_fecha_texto") or "—")

        banco = fila.get("banco")
        c_banco = ws2.cell(row=r, column=9, value=banco or "—")
        if fila.get("banco_rechazado"):
            c_banco.fill = ROJO

        # CLABE y número de cuenta se exportan como TEXTO (number_format "@")
        # para que Excel no los convierta a notación científica ni les
        # quite ceros a la izquierda.
        c_cta = ws2.cell(row=r, column=10, value=fila.get("numero_cuenta") or "—")
        c_cta.number_format = "@"
        c_clabe = ws2.cell(row=r, column=11, value=fila.get("clabe") or "—")
        c_clabe.number_format = "@"

        ws2.cell(row=r, column=12, value=fila["detalle"] + (" | Muestra: " + fila["texto_muestra"] if fila["legible"] else ""))
        ws2.cell(row=r, column=12).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    _autoancho(ws2, [30, 26, 9, 10, 12, 20, 12, 24, 22, 18, 20, 70])
    for fila_idx in range(2, r):
        ws2.row_dimensions[fila_idx].height = 45

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Valida el expediente de un candidato y genera un Excel.")
    ap.add_argument("--nombre", required=True, help="Nombre completo registrado del candidato")
    ap.add_argument("--rfc", default="", help="RFC capturado (opcional)")
    ap.add_argument("--curp", default="", help="CURP capturado (opcional)")
    ap.add_argument("--salida", default="expediente_validado.xlsx", help="Ruta del Excel de salida")
    ap.add_argument("pdfs", nargs="+", help="Rutas a los PDFs del candidato")
    args = ap.parse_args()

    candidato = {"nombre": args.nombre, "rfc": args.rfc, "curp": args.curp}

    filas = []
    for ruta in args.pdfs:
        print(f"Procesando: {os.path.basename(ruta)} ...")
        try:
            fila = procesar_documento(ruta, candidato["nombre"])
        except Exception as e:
            fila = {
                "archivo": os.path.basename(ruta), "categoria_clave": "ERROR",
                "categoria": "Error al procesar", "num_paginas": 0, "uso_ocr": False,
                "legible": False, "texto_muestra": "", "nombre_coincide": None,
                "detalle": f"Error: {e}",
            }
        filas.append(fila)

    checklist, extra = construir_reporte(candidato, filas)
    generar_excel(candidato, filas, checklist, extra, args.salida)
    print(f"\nListo. Reporte guardado en: {args.salida}")


if __name__ == "__main__":
    main()
